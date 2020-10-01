'''
============================
@Time    :2020/03/08/14:48
@Author  :cai ming liang
@E-mail  :2117899275@qq.com
============================
'''
import openpyxl

class ReadExcle(object):

    def __init__(self,file_name,sheet_name):
        self.file_name=file_name
        self.sheet_name=sheet_name

    def read_data(self):
        work_book=openpyxl.load_workbook(self.file_name)
        sheet=work_book[self.sheet_name]
        datas=list(sheet.rows)

        keys=[i.value for i in datas[0]]
        data_list=[]
        for i in datas[1:]:
            values=[j.value for j in i]
            case=dict(zip(keys,values))
            data_list.append(case)
        return data_list

    def write_data(self,row,column,value):
        work_book = openpyxl.load_workbook(self.file_name)
        sheet = work_book[self.sheet_name]
        sheet.cell(row=row,column=column,value=value)
        work_book.save(self.file_name)
